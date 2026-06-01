"""
Importa Emendas_Impositivas_2026.xlsx para o banco (emendas estaduais SP).

Mapeamento de colunas:
  0  MUNICÍPIO        → municipioNome + lookup IBGE
  1  ÓRGÃO PROCESSADOR → orgaoExecutor
  2  OBJETO           → objeto
  3  PARLAMENTAR      → parlamentar.nome (DEPUTADO_ESTADUAL)
  4  PARTIDO          → parlamentar.partido
  5  ANO              → ano
  6  CÓDIGO           → idPortal (= numero)
  7  Primeira Fase Substituídas → (ignorado)
  8  TIPO EMENDA      → tipo
  9  FUNÇÃO DE GOVERNO → funcao + area
 10  BENEFICIÁRIO     → beneficiario
 11  CNPJ BENEFICIÁRIO→ cnpjBeneficiario
 12  ESTÁGIO          → determina valorEmpenhado/valorPago
 13  DATA PAGAMENTO   → (ignorado)
 14  VALOR DECISÃO    → valor base
 15  VALOR REMANEJADO → valor alternativo quando DECISÃO ausente

Regras de valor:
  - valorProposto   = DECISÃO ?? REMANEJADO ?? 0  (todos os estágios)
  - valorEmpenhado  = valorProposto quando estágio in {Empenhado/Convênio, Pagas}
  - valorPago       = valorProposto quando estágio == Pagas
"""

import sys
import os
import re
import unicodedata

# Force UTF-8 stdout on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime

# ── Configuração ──────────────────────────────────────────────────────────────

XLSX = os.path.join(os.path.dirname(__file__), '..', 'data', 'estados',
                    'Emendas_Impositivas_2026.xlsx')

DATABASE_URL = (
    'postgresql://postgres.mtjugadcnwjbcpfzavxs:Meug4binete'
    '@aws-1-sa-east-1.pooler.supabase.com:5432/postgres'
)

UF = 'SP'
ESFERA = 'ESTADUAL'
CARGO = 'DEPUTADO_ESTADUAL'

ESTAGIOS_EMPENHADO = {'empenhado / convênio', 'pagas'}
ESTAGIOS_PAGO      = {'pagas'}

BATCH = 200

# Grafias divergentes entre a fonte e o banco (chave = sem_acento(UPPER(nome_excel)))
ALIAS_MUNICIPIO: dict[str, str] = {
    'LUIZ ANTONIO': 'LUIS ANTONIO',   # Excel usa Z, banco usa S
}

# ── Classificação de área ──────────────────────────────────────────────────────

FUNCAO_AREA = {
    '10': 'SAUDE',
    '12': 'HABITACAO',
    '08': 'ASSISTENCIA_SOCIAL',
    '09': 'ASSISTENCIA_SOCIAL',
    '11': 'EDUCACAO',
    '13': 'EDUCACAO',
    '26': 'INFRAESTRUTURA',
    '15': 'INFRAESTRUTURA',
    '17': 'INFRAESTRUTURA',
    '28': 'OUTROS',       # ENCARGOS_ESPECIAIS não existe no enum
    '06': 'SEGURANCA',
    '05': 'SEGURANCA',
    '18': 'INFRAESTRUTURA',  # GESTÃO AMBIENTAL
    '20': 'AGRICULTURA',
    '27': 'ESPORTE',
    '13': 'CULTURA',
    '16': 'HABITACAO',
    '22': 'SANEAMENTO',
    '04': 'INFRAESTRUTURA',  # ADMINISTRAÇÃO
    '14': 'SANEAMENTO',
}

def classificar_area(funcao: str | None) -> str:
    if not funcao:
        return 'OUTROS'
    m = re.match(r'^(\d+)', str(funcao).strip())
    if m:
        return FUNCAO_AREA.get(m.group(1), 'OUTROS')
    f = funcao.lower()
    if 'saúde' in f or 'saude' in f:          return 'SAUDE'
    if 'educação' in f or 'educac' in f:       return 'EDUCACAO'
    if 'saneamento' in f:                      return 'SANEAMENTO'
    if 'transp' in f or 'infraes' in f:        return 'INFRAESTRUTURA'
    if 'assistência' in f or 'social' in f:    return 'ASSISTENCIA_SOCIAL'
    if 'segurança' in f or 'segur' in f:       return 'SEGURANCA'
    if 'habitação' in f or 'habitac' in f:     return 'HABITACAO'
    if 'agricultur' in f or 'agrope' in f:     return 'AGRICULTURA'
    if 'esporte' in f or 'desport' in f:       return 'ESPORTE'
    if 'cultura' in f:                         return 'CULTURA'
    if 'meio ambiente' in f or 'ambient' in f: return 'MEIO_AMBIENTE'
    return 'OUTROS'

# ── Helpers ────────────────────────────────────────────────────────────────────

def sem_acento(s: str) -> str:
    """Remove diacríticos e normaliza hífen→espaço para comparação de nomes."""
    s = s.replace('-', ' ')
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

def to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def cuid_like() -> str:
    import random, string, time
    ts = hex(int(time.time() * 1000))[2:]
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=20))
    return 'c' + ts + rand

def normalizar_nome(s: str) -> str:
    """TÍTULO DE PALAVRAS normalizado."""
    if not s:
        return s
    stop = {'de', 'da', 'do', 'das', 'dos', 'e', 'em', 'na', 'no', 'nas', 'nos'}
    words = s.strip().lower().split()
    return ' '.join(
        w.capitalize() if i == 0 or w not in stop else w
        for i, w in enumerate(words)
    )

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f'[{datetime.now():%H:%M:%S}] Abrindo {XLSX}…')
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f'[{datetime.now():%H:%M:%S}] {len(rows)} linhas carregadas.')

    print(f'[{datetime.now():%H:%M:%S}] Conectando ao banco…')
    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # ── 1. Carregar mapa nome→IBGE dos municípios SP ──────────────────────────
    cur.execute("""
        SELECT DISTINCT ON (UPPER(TRIM(nome)))
               UPPER(TRIM(nome)) AS nome_upper,
               "codigoIbge"
          FROM municipio_stats
         WHERE uf = %s
         ORDER BY UPPER(TRIM(nome)), ano DESC
    """, (UF,))
    # Chave normalizada (sem acento, hífen→espaço) para tolerar divergências ortográficas
    ibge_map = {sem_acento(r['nome_upper']): r['codigoIbge'] for r in cur.fetchall()}
    print(f'[{datetime.now():%H:%M:%S}] {len(ibge_map)} municípios SP no banco.')

    # ── 2. Carregar parlamentares existentes (DEPUTADO_ESTADUAL) ─────────────
    cur.execute("""
        SELECT id, UPPER(TRIM(nome)) AS nome_upper
          FROM parlamentares
         WHERE cargo = 'DEPUTADO_ESTADUAL'
    """)
    parl_map: dict[str, str] = {r['nome_upper']: r['id'] for r in cur.fetchall()}
    print(f'[{datetime.now():%H:%M:%S}] {len(parl_map)} deputados estaduais já no banco.')

    # ── 3. Processar linhas ───────────────────────────────────────────────────
    muni_sem_ibge: set[str] = set()
    emendas_upsert = []
    parls_novos: dict[str, dict] = {}

    for row in rows:
        municipio_raw = str(row[0] or '').strip()
        orgao         = str(row[1] or '').strip() or None
        objeto        = str(row[2] or '').strip() or None
        parl_nome_raw = str(row[3] or '').strip()
        partido       = str(row[4] or '').strip() or None
        ano           = int(row[5]) if row[5] else 2026
        codigo        = str(row[6] or '').strip() or None
        tipo          = str(row[8] or '').strip() or None
        funcao        = str(row[9] or '').strip() or None
        beneficiario  = str(row[10] or '').strip() or None
        cnpj          = str(row[11] or '').strip() or None
        estagio       = str(row[12] or '').strip().lower()
        val_dec       = to_float(row[14])
        val_rem       = to_float(row[15])

        if not codigo or not parl_nome_raw:
            continue

        # Valor base: DECISÃO tem prioridade, senão REMANEJADO
        valor_base = val_dec if val_dec > 0 else val_rem

        valor_empenhado = valor_base if estagio in ESTAGIOS_EMPENHADO else 0.0
        valor_pago      = valor_base if estagio in ESTAGIOS_PAGO      else 0.0

        # IBGE lookup — normaliza acentos/hífen e aplica aliases de grafia divergente
        muni_upper = municipio_raw.upper()
        muni_key = sem_acento(muni_upper)
        muni_key = ALIAS_MUNICIPIO.get(muni_key, muni_key)
        codigo_ibge = ibge_map.get(muni_key)
        if not codigo_ibge and municipio_raw:
            muni_sem_ibge.add(municipio_raw)

        municipio_nome = normalizar_nome(municipio_raw) if municipio_raw else None
        area = classificar_area(funcao)

        # Parlamentar
        parl_upper = parl_nome_raw.upper()
        if parl_upper not in parl_map and parl_upper not in parls_novos:
            parls_novos[parl_upper] = {
                'id':      cuid_like(),
                'nome':    normalizar_nome(parl_nome_raw),
                'partido': partido,
                'uf':      UF,
                'cargo':   CARGO,
            }

        emendas_upsert.append({
            'codigo':         codigo,
            'esfera':         ESFERA,
            'ano':            ano,
            'numero':         codigo,
            'tipo':           tipo,
            'funcao':         funcao,
            'area':           area,
            'objeto':         objeto,
            'orgao':          orgao,
            'beneficiario':   beneficiario,
            'cnpj':           cnpj,
            'uf':             UF,
            'codigoIbge':     codigo_ibge,
            'municipioNome':  municipio_nome,
            'valEmp':         valor_empenhado,
            'valPago':        valor_pago,
            'valProposto':    valor_base,
            'parlUpper':      parl_upper,
        })

    print(f'[{datetime.now():%H:%M:%S}] {len(emendas_upsert)} emendas para upsert.')
    print(f'[{datetime.now():%H:%M:%S}] {len(parls_novos)} novos parlamentares.')
    if muni_sem_ibge:
        print(f'[{datetime.now():%H:%M:%S}] ⚠  {len(muni_sem_ibge)} municípios sem IBGE (sem coloração no mapa):')
        for m in sorted(muni_sem_ibge)[:20]:
            print(f'     - {m}')

    # ── 4. Inserir novos parlamentares ────────────────────────────────────────
    now = datetime.utcnow()
    for parl_upper, p in parls_novos.items():
        cur.execute("""
            INSERT INTO parlamentares (id, nome, partido, uf, cargo, ativo, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, %s, %s, true, %s, %s)
            ON CONFLICT DO NOTHING
        """, (p['id'], p['nome'], p['partido'], p['uf'], p['cargo'], now, now))
        parl_map[parl_upper] = p['id']
    conn.commit()
    print(f'[{datetime.now():%H:%M:%S}] Parlamentares inseridos.')

    # ── 5. Upsert emendas em batches (com reconexão automática) ──────────────
    INSERT_SQL = """
        INSERT INTO emendas_parlamentares (
            id, "idPortal", esfera, ano, numero, tipo, funcao, area,
            objeto, "orgaoExecutor", beneficiario, "cnpjBeneficiario",
            uf, "codigoIbge", "municipioNome",
            "valorEmpenhado", "valorPago", "valorProposto", "valorRestoPago",
            "parlamentarId", "fetchedAt", "updatedAt"
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, 0,
            %s, %s, %s
        )
        ON CONFLICT ("idPortal") DO UPDATE SET
            esfera          = EXCLUDED.esfera,
            ano             = EXCLUDED.ano,
            tipo            = EXCLUDED.tipo,
            funcao          = EXCLUDED.funcao,
            area            = EXCLUDED.area,
            objeto          = EXCLUDED.objeto,
            "orgaoExecutor" = EXCLUDED."orgaoExecutor",
            beneficiario    = EXCLUDED.beneficiario,
            "cnpjBeneficiario" = EXCLUDED."cnpjBeneficiario",
            uf              = EXCLUDED.uf,
            "codigoIbge"    = EXCLUDED."codigoIbge",
            "municipioNome" = EXCLUDED."municipioNome",
            "valorEmpenhado"= EXCLUDED."valorEmpenhado",
            "valorPago"     = EXCLUDED."valorPago",
            "valorProposto" = EXCLUDED."valorProposto",
            "parlamentarId" = EXCLUDED."parlamentarId",
            "updatedAt"     = EXCLUDED."updatedAt"
    """

    def reconectar():
        c = psycopg2.connect(DATABASE_URL)
        return c, c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    total = len(emendas_upsert)
    inseridas = 0
    for i in range(0, total, BATCH):
        batch = emendas_upsert[i:i+BATCH]
        tentativas = 0
        while True:
            try:
                for e in batch:
                    parl_id = parl_map.get(e['parlUpper'])
                    cur.execute(INSERT_SQL, (
                        cuid_like(),
                        e['codigo'], e['esfera'], e['ano'], e['numero'], e['tipo'], e['funcao'], e['area'],
                        e['objeto'], e['orgao'], e['beneficiario'], e['cnpj'],
                        e['uf'], e['codigoIbge'], e['municipioNome'],
                        e['valEmp'], e['valPago'], e['valProposto'],
                        parl_id, now, now,
                    ))
                conn.commit()
                inseridas += len(batch)
                break
            except psycopg2.OperationalError as ex:
                tentativas += 1
                if tentativas > 3:
                    raise
                print(f'\n[{datetime.now():%H:%M:%S}] Conexão perdida ({ex}), reconectando ({tentativas}/3)…')
                try: conn.close()
                except Exception: pass
                conn, cur = reconectar()

        pct = inseridas * 100 // total
        print(f'[{datetime.now():%H:%M:%S}] {inseridas}/{total} ({pct}%)…', end='\r', flush=True)

    print(f'\n[{datetime.now():%H:%M:%S}] ✅ {inseridas} emendas importadas com sucesso.')
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
