"""
Importa Emendas Impositivas SP — todos os anos disponíveis.

Fontes:
  XLSX (2023–2026): Emendas_Impositivas_{ANO}.xlsx
    Colunas: 0=MUNICÍPIO 1=ÓRGÃO 2=OBJETO 3=PARLAMENTAR 4=PARTIDO
             5=ANO 6=CÓDIGO 7=PrimeiraFase 8=TIPO 9=FUNÇÃO
            10=BENEFICIÁRIO 11=CNPJ 12=ESTÁGIO 13=DATA 14=VALOR_DECISÃO 15=VALOR_REMANEJADO

  PDF (2021–2022): {ANO}_Saude.pdf + {ANO}_ExcetoSaude.pdf
    Colunas: 0=PARLAMENTAR 1=BENEFICIÁRIO 2=MUNICÍPIO 3=OBJETO 4=ÓRGÃO 5=VALOR 6=STATUS

Regras de valor (xlsx):
  valorProposto   = VALOR_DECISÃO ?? VALOR_REMANEJADO
  valorEmpenhado  = valorProposto  (para todos os estágios — valor aprovado/destinado)
  valorPago       = valorProposto  somente quando estágio == "pagas"

Regras de valor (PDF):
  valorProposto = valorEmpenhado = VALOR (R$ X.XXX,XX)
  valorPago     = VALOR somente quando status == "PAGA"
"""

import sys, os, re, unicodedata
import openpyxl, pdfplumber
import psycopg2, psycopg2.extras
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# ── Configuração ───────────────────────────────────────────────────────────────

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'estados')

DATABASE_URL = (
    'postgresql://postgres.mtjugadcnwjbcpfzavxs:Meug4binete'
    '@aws-1-sa-east-1.pooler.supabase.com:5432/postgres'
)

UF     = 'SP'
ESFERA = 'ESTADUAL'
CARGO  = 'DEPUTADO_ESTADUAL'
BATCH  = 200

ALIAS_MUNICIPIO: dict[str, str] = {
    'LUIZ ANTONIO': 'LUIS ANTONIO',
}

# ── Área ───────────────────────────────────────────────────────────────────────

FUNCAO_AREA = {
    '10': 'SAUDE',   '08': 'ASSISTENCIA_SOCIAL', '09': 'ASSISTENCIA_SOCIAL',
    '11': 'EDUCACAO','13': 'CULTURA',             '12': 'HABITACAO',
    '16': 'HABITACAO','26': 'INFRAESTRUTURA',     '15': 'INFRAESTRUTURA',
    '17': 'INFRAESTRUTURA','18': 'MEIO_AMBIENTE',  '04': 'INFRAESTRUTURA',
    '14': 'SANEAMENTO','22': 'SANEAMENTO',         '06': 'SEGURANCA',
    '05': 'SEGURANCA','20': 'AGRICULTURA',         '27': 'ESPORTE',
    '28': 'OUTROS',
}

ORGAO_AREA_KW = [
    ('SAÚDE', 'SAUDE'), ('SAUDE', 'SAUDE'),
    ('EDUCAÇÃO', 'EDUCACAO'), ('EDUCACAO', 'EDUCACAO'), ('ESCOLA', 'EDUCACAO'),
    ('SOCIAL', 'ASSISTENCIA_SOCIAL'), ('CIDADANIA', 'ASSISTENCIA_SOCIAL'),
    ('ESPORTE', 'ESPORTE'), ('ESPORTES', 'ESPORTE'),
    ('CULTURA', 'CULTURA'),
    ('SEGURANÇA', 'SEGURANCA'), ('SEGURANCA', 'SEGURANCA'), ('BOMBEIROS', 'SEGURANCA'),
    ('AGRICULTURA', 'AGRICULTURA'), ('ANIMAL', 'AGRICULTURA'),
    ('HABITAÇÃO', 'HABITACAO'), ('HABITACAO', 'HABITACAO'),
    ('SANEAMENTO', 'SANEAMENTO'),
    ('TRANSPORT', 'INFRAESTRUTURA'), ('OBRAS', 'INFRAESTRUTURA'), ('URBAN', 'INFRAESTRUTURA'),
    ('MEIO AMBIENTE', 'MEIO_AMBIENTE'), ('AMBIENTAL', 'MEIO_AMBIENTE'),
]

def sem_acento(s: str) -> str:
    s = s.replace('-', ' ')
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def area_de_funcao(funcao: str | None) -> str:
    if not funcao:
        return 'OUTROS'
    m = re.match(r'^(\d+)', str(funcao).strip())
    if m:
        return FUNCAO_AREA.get(m.group(1), 'OUTROS')
    return area_de_orgao(funcao)

def area_de_orgao(orgao: str, saude_pdf: bool = False) -> str:
    if saude_pdf:
        return 'SAUDE'
    o = sem_acento(orgao.upper())
    for kw, area in ORGAO_AREA_KW:
        if sem_acento(kw) in o:
            return area
    return 'OUTROS'

# ── Helpers ────────────────────────────────────────────────────────────────────

def normalizar_nome(s: str) -> str:
    if not s:
        return s
    stop = {'de', 'da', 'do', 'das', 'dos', 'e', 'em', 'na', 'no', 'nas', 'nos'}
    words = s.strip().lower().split()
    return ' '.join(w.capitalize() if i == 0 or w not in stop else w for i, w in enumerate(words))

def to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        pass
    s = str(v).strip()
    s = re.sub(r'[R$\s]', '', s).replace('.', '').replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def cuid_like() -> str:
    import random, string, time
    ts   = hex(int(time.time() * 1000))[2:]
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=20))
    return 'c' + ts + rand

def ibge_lookup(municipio_raw: str, ibge_map: dict) -> tuple[str | None, str | None]:
    if not municipio_raw:
        return None, None
    key = sem_acento(municipio_raw.upper())
    key = ALIAS_MUNICIPIO.get(key, key)
    codigo = ibge_map.get(key)
    return normalizar_nome(municipio_raw), codigo

# ── Parse xlsx (2023-2026) ────────────────────────────────────────────────────

def parse_xlsx(ano: int, ibge_map: dict) -> list[dict]:
    path = os.path.join(DATA_DIR, f'Emendas_Impositivas_{ano}.xlsx')
    if not os.path.exists(path):
        print(f'  {ano}: arquivo não encontrado, pulando.')
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    next(ws.iter_rows(min_row=1, max_row=1))  # skip header

    emendas = []
    sem_ibge = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        municipio_raw = str(row[0] or '').strip()
        orgao         = str(row[1] or '').strip() or None
        objeto        = str(row[2] or '').strip() or None
        parl_raw      = str(row[3] or '').strip()
        partido       = str(row[4] or '').strip() or None
        codigo        = str(row[6] or '').strip() or None
        tipo          = str(row[8] or '').strip() or None
        funcao        = str(row[9] or '').strip() or None
        beneficiario  = str(row[10] or '').strip() or None
        cnpj          = str(row[11] or '').strip() or None
        estagio       = str(row[12] or '').strip().lower()
        val_dec       = to_float(row[14])
        val_rem       = to_float(row[15])

        if not codigo or not parl_raw:
            continue

        valor_base = val_dec if val_dec > 0 else val_rem
        if valor_base == 0:
            continue

        muni_nome, codigo_ibge = ibge_lookup(municipio_raw, ibge_map)
        if not codigo_ibge and municipio_raw:
            sem_ibge += 1

        # Valor: aprovado para todos; pago apenas quando executado
        val_pago = valor_base if estagio == 'pagas' else 0.0

        emendas.append({
            'idPortal':      codigo,
            'numero':        codigo,
            'ano':           ano,
            'tipo':          tipo,
            'funcao':        funcao or orgao,
            'area':          area_de_funcao(funcao) if funcao else area_de_orgao(orgao or ''),
            'objeto':        objeto[:500] if objeto else None,
            'orgao':         orgao[:200] if orgao else None,
            'beneficiario':  beneficiario[:300] if beneficiario else None,
            'cnpj':          cnpj[:20] if cnpj else None,
            'codigoIbge':    codigo_ibge,
            'municipioNome': muni_nome,
            'valEmp':        valor_base,
            'valPago':       val_pago,
            'valProp':       valor_base,
            'partido':       partido,
            'parlUpper':     parl_raw.upper(),
            'parlNome':      normalizar_nome(parl_raw),
        })

    wb.close()
    print(f'  {ano} xlsx: {len(emendas)} emendas | sem IBGE: {sem_ibge}')
    return emendas

# ── Parse PDF (2021-2022) ─────────────────────────────────────────────────────

def parse_pdf(ano: int, fname: str, ibge_map: dict, saude: bool) -> list[dict]:
    path = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path):
        print(f'  {fname}: não encontrado, pulando.')
        return []

    emendas = []
    sem_ibge = 0
    idx = 0

    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            for table in pg.extract_tables():
                for row in table:
                    if not row or not row[0]:
                        continue
                    parl_raw = str(row[0]).replace('\n', ' ').strip()
                    if not parl_raw or parl_raw in ('PARLAMENTAR',) or 'EMENDAS IMPOSITIVAS' in parl_raw:
                        continue
                    if len(row) < 6 or not row[5] or 'R$' not in str(row[5]):
                        continue

                    beneficiario  = str(row[1] or '').replace('\n', ' ').strip() or None
                    municipio_raw = str(row[2] or '').replace('\n', ' ').strip()
                    objeto        = str(row[3] or '').replace('\n', ' ').strip() or None
                    orgao         = str(row[4] or '').replace('\n', ' ').strip() or None
                    status        = str(row[6] or '').replace('\n', ' ').strip().upper() if len(row) > 6 else ''

                    valor = to_float(row[5])
                    if valor == 0:
                        continue

                    muni_nome, codigo_ibge = ibge_lookup(municipio_raw, ibge_map)
                    if not codigo_ibge and municipio_raw:
                        sem_ibge += 1

                    val_pago = valor if 'PAGA' in status else 0.0
                    prefix   = 'saude' if saude else 'outros'

                    emendas.append({
                        'idPortal':      f'SP{ano}_{prefix}_r{idx}',
                        'numero':        None,
                        'ano':           ano,
                        'tipo':          'EMENDA LOA',
                        'funcao':        orgao[:100] if orgao else None,
                        'area':          area_de_orgao(orgao or '', saude_pdf=saude),
                        'objeto':        objeto[:500] if objeto else None,
                        'orgao':         orgao[:200] if orgao else None,
                        'beneficiario':  beneficiario[:300] if beneficiario else None,
                        'cnpj':          None,
                        'codigoIbge':    codigo_ibge,
                        'municipioNome': muni_nome,
                        'valEmp':        valor,
                        'valPago':       val_pago,
                        'valProp':       valor,
                        'partido':       None,
                        'parlUpper':     parl_raw.upper(),
                        'parlNome':      normalizar_nome(parl_raw),
                    })
                    idx += 1

    print(f'  {fname}: {len(emendas)} emendas | sem IBGE: {sem_ibge}')
    return emendas

# ── Upsert ─────────────────────────────────────────────────────────────────────

INSERT_SQL = """
    INSERT INTO emendas_parlamentares (
        id, "idPortal", esfera, ano, numero, tipo, funcao, area,
        objeto, "orgaoExecutor", beneficiario, "cnpjBeneficiario",
        uf, "codigoIbge", "municipioNome",
        "valorEmpenhado", "valorPago", "valorProposto", "valorRestoPago",
        "parlamentarId", "fetchedAt", "updatedAt"
    ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s,
        %s, %s, %s,
        %s, %s, %s, 0,
        %s, %s, %s
    )
    ON CONFLICT ("idPortal") DO UPDATE SET
        esfera              = EXCLUDED.esfera,
        ano                 = EXCLUDED.ano,
        tipo                = EXCLUDED.tipo,
        funcao              = EXCLUDED.funcao,
        area                = EXCLUDED.area,
        objeto              = EXCLUDED.objeto,
        "orgaoExecutor"     = EXCLUDED."orgaoExecutor",
        beneficiario        = EXCLUDED.beneficiario,
        "cnpjBeneficiario"  = EXCLUDED."cnpjBeneficiario",
        uf                  = EXCLUDED.uf,
        "codigoIbge"        = EXCLUDED."codigoIbge",
        "municipioNome"     = EXCLUDED."municipioNome",
        "valorEmpenhado"    = EXCLUDED."valorEmpenhado",
        "valorPago"         = EXCLUDED."valorPago",
        "valorProposto"     = EXCLUDED."valorProposto",
        "parlamentarId"     = EXCLUDED."parlamentarId",
        "updatedAt"         = EXCLUDED."updatedAt"
"""

def reconectar():
    c = psycopg2.connect(DATABASE_URL)
    return c, c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f'[{datetime.now():%H:%M:%S}] Conectando ao banco…')
    conn, cur = reconectar()

    cur.execute("""
        SELECT DISTINCT ON (UPPER(TRIM(nome)))
               UPPER(TRIM(nome)) AS nome_upper, "codigoIbge"
          FROM municipio_stats WHERE uf = 'SP'
         ORDER BY UPPER(TRIM(nome))
    """)
    ibge_map = {sem_acento(r['nome_upper']): r['codigoIbge'] for r in cur.fetchall()}
    print(f'[{datetime.now():%H:%M:%S}] {len(ibge_map)} municípios SP no banco.')

    cur.execute("""
        SELECT id, UPPER(TRIM(nome)) AS nome_upper, partido
          FROM parlamentares WHERE cargo = %s AND uf = %s
    """, (CARGO, UF))
    parl_rows = cur.fetchall()
    parl_map: dict[str, str] = {r['nome_upper']: r['id'] for r in parl_rows}
    print(f'[{datetime.now():%H:%M:%S}] {len(parl_map)} deputados SP já no banco.')

    now = datetime.utcnow()

    # ── 1. Coletar todas as emendas ────────────────────────────────────────────
    print(f'\n[{datetime.now():%H:%M:%S}] Parsing arquivos…')
    todas: list[dict] = []

    # PDF 2021
    todas.extend(parse_pdf(2021, '2021_Saude.pdf',       ibge_map, saude=True))
    todas.extend(parse_pdf(2021, '2021_ExcetoSaude.pdf', ibge_map, saude=False))
    # PDF 2022
    todas.extend(parse_pdf(2022, '2022_Saude.pdf',       ibge_map, saude=True))
    todas.extend(parse_pdf(2022, '2022_ExcetoSaude.pdf', ibge_map, saude=False))
    # XLSX 2023–2026
    for ano in [2023, 2024, 2025, 2026]:
        todas.extend(parse_xlsx(ano, ibge_map))

    print(f'\n[{datetime.now():%H:%M:%S}] Total: {len(todas)} emendas.')

    # ── 2. Novos parlamentares ─────────────────────────────────────────────────
    parls_novos: dict[str, dict] = {}
    for e in todas:
        pu = e['parlUpper']
        if pu not in parl_map and pu not in parls_novos:
            parls_novos[pu] = {
                'id':      cuid_like(),
                'nome':    e['parlNome'],
                'partido': e.get('partido'),
                'uf':      UF,
                'cargo':   CARGO,
            }

    print(f'[{datetime.now():%H:%M:%S}] {len(parls_novos)} novos parlamentares.')
    for pu, p in parls_novos.items():
        cur.execute("""
            INSERT INTO parlamentares (id, nome, partido, uf, cargo, ativo, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, %s, %s, true, %s, %s)
            ON CONFLICT DO NOTHING
        """, (p['id'], p['nome'], p['partido'], p['uf'], p['cargo'], now, now))
        parl_map[pu] = p['id']

    # Atualiza partido de parlamentares existentes quando disponível
    for e in todas:
        if e.get('partido') and e['parlUpper'] in parl_map:
            cur.execute("""
                UPDATE parlamentares SET partido = %s, "updatedAt" = %s
                 WHERE id = %s AND (partido IS NULL OR partido = '')
            """, (e['partido'], now, parl_map[e['parlUpper']]))
    conn.commit()
    print(f'[{datetime.now():%H:%M:%S}] Parlamentares ok.')

    # ── 3. Upsert emendas ─────────────────────────────────────────────────────
    total     = len(todas)
    inseridas = 0
    for i in range(0, total, BATCH):
        batch = todas[i:i + BATCH]
        tentativas = 0
        while True:
            try:
                for e in batch:
                    parl_id = parl_map.get(e['parlUpper'])
                    cur.execute(INSERT_SQL, (
                        cuid_like(),
                        e['idPortal'], ESFERA, e['ano'], e['numero'],
                        e['tipo'], e['funcao'], e['area'],
                        e['objeto'], e['orgao'], e['beneficiario'], e['cnpj'],
                        UF, e['codigoIbge'], e['municipioNome'],
                        e['valEmp'], e['valPago'], e['valProp'],
                        parl_id, now, now,
                    ))
                conn.commit()
                inseridas += len(batch)
                break
            except psycopg2.OperationalError:
                tentativas += 1
                if tentativas > 3:
                    raise
                print(f'\n  Reconectando ({tentativas}/3)…')
                try: conn.close()
                except Exception: pass
                conn, cur = reconectar()

        pct = inseridas * 100 // total
        print(f'  {inseridas}/{total} ({pct}%)…', end='\r', flush=True)

    print(f'\n[{datetime.now():%H:%M:%S}] ✅ {inseridas} emendas SP importadas/atualizadas.')
    cur.close()
    conn.close()


if __name__ == '__main__':
    main()
